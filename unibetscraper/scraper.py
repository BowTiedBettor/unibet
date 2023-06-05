import pandas as pd
from datetime import datetime, timedelta
from curl_cffi import requests as cffi_requests
from traceback import print_exc
from openpyxl import load_workbook
import smtplib
import json
import time
import sys

# insert email details for the sender address below, or load from local .env file
EMAIL_ADDRESS = "ADD_EMAIL_ADDRESS"
EMAIL_PASS = "ADD_EMAIL_PASS"

class UnibetScraper:
    def __init__(self, track, countrycode = "GBR", harness = False):
        """
        Generates the class object & defines the request headers for upcoming requests
        Stores the track name without åäö-letters in case those are included

        :param track str: Name of the track
        :param countrycode str: Three-letter code for the relevant country [can be found in the URL's for a race], 
                                defaults to "GBR" for British racing
        """
        # headers for the http requests [they stay the same for all requests]
        self.headers = {
                        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/96.0.4664.45 Safari/537.36',
                        'Accept': '*/*',
                        'Accept-Language': 'en-GB,en;q=0.5',
                        'Referer': 'https://www.unibet.co.uk/',
                        'content-type': 'application/json',
                        'Origin': 'https://www.unibet.co.uk',
                        'Connection': 'keep-alive',
                        'Sec-Fetch-Dest': 'empty',
                        'Sec-Fetch-Mode': 'cors',
                        'Sec-Fetch-Site': 'cross-site',
                    }
        
        # fix the track name if åäö is included
        letters = ["Å", "Ä", "Ö", "å", "ä", "ö"]
        adj_letters = ["A", "A", "O", "a", "a", "o"]
        adj_track = track
        for letter, adj_letter in zip(letters, adj_letters):
            adj_track = adj_track.replace(letter, adj_letter)
        
        self.track = adj_track
        self.countrycode = countrycode
        self.harness = harness

    def get_meeting(self, races: list):
        """
        Finds the races for the relevant meeting, collect & returns the
        the different eventkeys for each race that is to be scraped

        :param races list: A list provided as [start_race, until_race]

        :rtype: list of eventkeys
        """
        # apparently a maximum of 4 days diff or the server won't respond correctly
        # adjust the timedeltas in case of consecutive events at the same racetrack
        start = (datetime.now() - timedelta(days=1)).strftime('%Y-%m-%d')
        until = (datetime.now() + timedelta(days=2)).strftime('%Y-%m-%d') # adjust this for consecutive days at the same racetrack

        try:
            if self.harness: 
                response = cffi_requests.get(f"https://rsa.unibet.co.uk/api/v1/graphql?operationName=MeetingsByDateRange&variables=%7B%22startDateTime%22%3A%22{start}T23%3A00%3A00.000Z%22%2C%22endDateTime%22%3A%22{until}T23%3A00%3A00.000Z%22%2C%22countryCodes%22%3A%22{self.countrycode}%22%2C%22raceTypes%22%3A%5B%22H%22%5D%7D&extensions=%7B%22persistedQuery%22%3A%7B%22version%22%3A1%2C%22sha256Hash%22%3A%22b975a015a4d4e4dc298ebd6dd43e988ebf03fb940e2bb719478b59bde7bbffd6%22%7D%7D",
                    headers=self.headers)
            else: 
                response = cffi_requests.get(f"https://rsa.unibet.co.uk/api/v1/graphql?operationName=MeetingsByDateRange&variables=%7B%22startDateTime%22%3A%22{start}T23%3A00%3A00.000Z%22%2C%22endDateTime%22%3A%22{until}T23%3A00%3A00.000Z%22%2C%22countryCodes%22%3A%22{self.countrycode}%22%2C%22raceTypes%22%3A%5B%22T%22%5D%7D&extensions=%7B%22persistedQuery%22%3A%7B%22version%22%3A1%2C%22sha256Hash%22%3A%22b975a015a4d4e4dc298ebd6dd43e988ebf03fb940e2bb719478b59bde7bbffd6%22%7D%7D",
                    headers=self.headers)
            
            meetings = response.json()['data']['viewer']['meetingsByDateRange']

            # uncomment the below lines if you'd like to analyze the generated json data for a given race
            # with open('comps.json', 'w') as file:
            #     json.dump(meetings, file)

            for meet in meetings:
              if meet['name'] == self.track:
                events = meet['events']
                break

            eventkeys = []
            for event in events:
              racenr = int(event['eventKey'].split(".")[-1])
              if racenr >= races[0] and racenr <= races[1]:
                # saves the eventkeys for the relevant races
                eventkeys.append(event['eventKey'])

        except UnboundLocalError:
            # raised if no matching meeting could be found on the sportsbook
            # since the variable <events> isn't assigned anything
            print("No meeting for this racetrack could be found...")
            sys.exit()

        except KeyError as e:
            if 'data' in str(e):
                print("There was a problem with the get_meeting call, retrying in 1 sec...")
                time.sleep(1)
                self.get_meeting(races)

        except:
            print("Request failed...")
            print("More information below:")
            print_exc()
            sys.exit()

        return eventkeys

    def scrape_wp(self, races: list):
        """
        Scrapes the win & place markets

        :param race list: A list provided as [start_race, until_race]

        :rtype: list of pd.DataFrames
        """
        # calls the get_meeting method to obtain the eventkeys for the races
        eventkeys = self.get_meeting(races = races)

        # sets up a requests Session object to avoid having to establish a new connection for
        # every race
        session = cffi_requests.Session()
        session.headers = self.headers

        # creates an empty list that will hold the pandas dataframes [1 per race]
        pd_list = []

        for eventkey in eventkeys:
            # creates an empty dataframe for the race
            race_df = pd.DataFrame(
                columns=["PostPos", "Horse", "WOdds", "POdds"])
            try:
                response = session.get(f'https://rsa.unibet.co.uk/api/v1/graphql?operationName=EventQuery&variables=%7B%22clientCountryCode%22%3A%22SE%22%2C%22eventKey%22%3A%22{eventkey}%22%2C%22fetchTRC%22%3Afalse%7D&extensions=%7B%22persistedQuery%22%3A%7B%22version%22%3A1%2C%22sha256Hash%22%3A%228fee3aa36e812c03f4cb039cb2f2c2defc919ea7a89b9b942e5e79588260a8b7%22%7D%7D')

                competitors = response.json()['data']['viewer']['event']['competitors']

                # uncomment the below lines if you'd like to analyze the generated json data for a given race
                # with open('comps.json', 'w') as file:
                #     json.dump(competitors, file)

                for comp in competitors:
                    post_pos = comp['sequence']
                    name = comp['name']
                    prices = comp['prices']

                    if comp['status'] == "Scratched":
                        win_odds = "N/A"
                        place_odds = "N/A"

                    else:
                        win_odds = 0.00
                        place_odds = 0.00
                        if prices:
                            for price in prices:
                                if price['betType'] == "FixedWin":
                                    win_odds = round(price['price'], 2)
                                if price['betType'] == "FixedPlace":
                                    place_odds = round(price['price'], 2)

                    dummy_df = race_df
                    new_row = pd.DataFrame(
                        [[post_pos, name, win_odds, place_odds]], columns=race_df.columns)
                    race_df = pd.concat(
                        [dummy_df, new_row], ignore_index=True)
            except:
                # prints the exception and continues with the next race instead
                print_exc()
                continue

            # when the race has been completed, sort on post position and append to pd_list
            sorted_df = race_df.sort_values(by="PostPos")
            pd_list.append(sorted_df)

        return pd_list

    def scrape_h2h(self, races: list):
        """
        Scrapes the H2H markets

        :param races list: A list provided as [start_race, until_race]

        :rtype: list of pd.DataFrames
        """
        params = {
                'operationName': 'SpecialsQuery',
                'variables': '{"locationClass":"Special"}',
                'extensions': '{"persistedQuery":{"version":1,"sha256Hash":"5c9c77d8d13e860ac8cf5ca74eab13d9039bbfbbbfe56db12d2d22244c90c65f"}}',
            }

        while True:
            try:
                response_json = cffi_requests.get('https://rsa.unibet.co.uk/api/v1/graphql', params=params, headers=self.headers).json()

                specials = response_json['data']['viewer']['specials']
                break
            except:
                print(f"Problem with the 'specials' call [H2H], retrying in 1 sec...")
                time.sleep(1)

        # creates an empty list that will hold the pandas dataframes [1 per h2h-object]
        pd_list = []

        for special in specials:
            if self.track in special['name']: # add check for correct racenr
                h2h_selections = special['competitors']

                # creates an empty dataframe for the h2h object
                h2h_df = pd.DataFrame(
                    columns=["PostPos", "Horse", "H2H-odds"])

                for comp in h2h_selections:
                    postpos = comp['sequence']
                    name = comp['name']
                    prices = comp['prices']

                    h2h_odds = 0.00
                    if prices:
                        for price in prices:
                            if price['betType'] == "FixedWin":
                                h2h_odds = round(price['price'], 2)

                    dummy_df = h2h_df
                    ny_rad = pd.DataFrame(
                        [[postpos, name, h2h_odds]], columns=h2h_df.columns)
                    h2h_df = pd.concat(
                        [dummy_df, ny_rad], ignore_index=True)

                # when the race has been completed, sort on post position and append to pd_list
                sorted_df = h2h_df.sort_values(by="PostPos")
                pd_list.append(sorted_df)

        return pd_list

    def awaitnewodds(self, races: list, delta: int):
        """
        """
        # calls the get_meeting method to obtain the eventkeys for the races
        eventkeys = self.get_meeting(races = races)

        # queries the API to get info for all the races
        params = {
                'operationName': 'MultiCouponEventsQuery',
                'variables': json.dumps({"eventKeys": eventkeys}),
                'extensions': '{"persistedQuery":{"version":1,"sha256Hash":"b8c95b2e242eee16704457a7f0b427d01b558b8cb261292efcadb84b5ea987bb"}}',
            }

        while True:
            try:
                response_json = cffi_requests.get('https://rsa.unibet.co.uk/api/v1/graphql', params=params, headers=self.headers).json()

                events = response_json['data']['viewer']['events']
                # loops through the events & checks whether the status for any of them is "Open"
                for event in events:
                    if event['status'] == "Open":
                        print(f"Unibet - Odds for {self.track} have now been released!")
                        return True
                # if no event status was "Open", an exception is raised to execute the except block
                # note that if there was a problem with the request then it retries
                raise Exception

            except:
                print(datetime.now())
                print("Unibet")
                print("No odds were found")
                print(f"Retries in {delta} seconds...")
                print()
                time.sleep(delta)
                continue

    def send_mail(self, recipients: list, result: list):
        """
        Sends an email containing all the information to the email addresses provided in the <recipients> list

        :param recipients list: List with email addresses that are to be notified with the scraped information
        :param result list: List of pandas dataFrames to be sent

        :rtype: None
        """
        with smtplib.SMTP('smtp.gmail.com', 587) as smtp:
            smtp.ehlo()
            smtp.starttls()
            smtp.ehlo()
            smtp.login(EMAIL_ADDRESS, EMAIL_PASS)

            subject = "UNIBET ODDS"
            body = f"Unibet odds for {self.track}: \n\n"
            for res in result:
                body += str(res) + "\n" + "\n"

            msg = f"Subject: {subject}\n\n{body}".encode('utf-8')

            for recipient in recipients:
                smtp.sendmail(EMAIL_ADDRESS, recipient, msg)
                print(f"Mail has been delivered to {recipient}.")

    def to_excel(self, pd_list: list, pd_list_name: str, file_path: str, sheet_name: str):
        """
        Dumps a given set of scraped races onto an Excel file, appends to previous information
        by storing the new data in new columns

        :param pd_list list: List of pandas dataFrames, one DF for each race
        :param pd_list_name str: A string containing the variable name to clarify any failures in output 
                                [in case list is empty, common for H2H calls]
        :param file_path str: File path for the Excel file
        :param sheet_name str: Name of the sheet the data is supposed to be stored in

        :rtype: None
        """
        if not pd_list:
            # if pd_list is empty then there's nothing to push to excel
            print(f"The list {pd_list_name} of dataframes is empty, nothing to append to Excel...")
            print()

        else:
            current_col = pd.read_excel(file_path, sheet_name = sheet_name).shape[1]
            time = datetime.now().strftime('%H:%M:%S')

            # insert the time the odds was pulled in the correct cell
            workbook = load_workbook(file_path)
            worksheet = workbook[sheet_name]
            if current_col == 0:
                worksheet.cell(row=1, column=2).value = time
            else:
                worksheet.cell(row=1, column = current_col + 4).value = time
            workbook.save(file_path)

            # insert the scraped odds
            i = 0
            for racedf in pd_list:
                with pd.ExcelWriter(file_path, mode='a', if_sheet_exists='overlay', engine = 'openpyxl') as writer:
                    if current_col == 0:
                        if "W&P" in sheet_name:
                            racedf.to_excel(writer, sheet_name=sheet_name, startrow=1 + 18 * i, startcol=0, index = False)
                            i += 1
                        elif "H2H" in sheet_name:
                            racedf.to_excel(writer, sheet_name=sheet_name, startrow=1 + 5 * i, startcol=0, index = False)
                            i += 1
                    else:
                        if "W&P" in sheet_name:
                            racedf.to_excel(writer, sheet_name=sheet_name, startrow=1 + 18 * i, startcol=current_col + 2, index = False)
                            i += 1
                        elif "H2H" in sheet_name:
                            racedf.to_excel(writer, sheet_name=sheet_name, startrow=1 + 5 * i, startcol=current_col + 2, index = False)
                            i += 1
            print(f"{pd_list_name} mapped to Excel...")
            print()

    def get_historical_odds(self, race: int, horse: str):
        """
        """
        pass

    def plot_historical_odds(self, race: int, horse: str): 
        """
        """
        pass
